import os
import openpyxl
import json

excel_path = 'dichvucong_xay_dung_crawled_2026-07-17/phan-loai-tthc-cho-ai-agent.xlsx'
output_dir = 'src/data'
os.makedirs(output_dir, exist_ok=True)

wb = openpyxl.load_workbook(excel_path)

# 1. Export Phan_loai_TTHC
ws = wb['Phan_loai_TTHC']
headers = [str(c.value) for c in ws[1]]
phan_loai_data = []
for row_idx in range(2, ws.max_row + 1):
    row_data = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=row_idx, column=col_idx).value
        if val is not None:
            row_data[headers[col_idx-1]] = val
    if row_data:
        phan_loai_data.append(row_data)

with open(os.path.join(output_dir, 'phan_loai_tthc.json'), 'w', encoding='utf-8') as f:
    json.dump(phan_loai_data, f, ensure_ascii=False, indent=2)
print(f"Exported Phan_loai_TTHC: {len(phan_loai_data)} rows")

# 2. Export Cay_hoi_AI
ws = wb['Cay_hoi_AI']
headers = [str(c.value) for c in ws[1]]
cay_hoi_data = []
for row_idx in range(2, ws.max_row + 1):
    row_data = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=row_idx, column=col_idx).value
        if val is not None:
            row_data[headers[col_idx-1]] = val
    if row_data:
        cay_hoi_data.append(row_data)

with open(os.path.join(output_dir, 'cay_hoi_ai.json'), 'w', encoding='utf-8') as f:
    json.dump(cay_hoi_data, f, ensure_ascii=False, indent=2)
print(f"Exported Cay_hoi_AI: {len(cay_hoi_data)} rows")

# 3. Export Cau_hoi_theo_de_muc
ws = wb['Cau_hoi_theo_de_muc']
headers = [str(c.value) for c in ws[1]]
cau_hoi_data = []
for row_idx in range(2, ws.max_row + 1):
    row_data = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=row_idx, column=col_idx).value
        if val is not None:
            row_data[headers[col_idx-1]] = val
    if row_data:
        cau_hoi_data.append(row_data)

with open(os.path.join(output_dir, 'cau_hoi_theo_de_muc.json'), 'w', encoding='utf-8') as f:
    json.dump(cau_hoi_data, f, ensure_ascii=False, indent=2)
print(f"Exported Cau_hoi_theo_de_muc: {len(cau_hoi_data)} rows")
